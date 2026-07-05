"""
database/db.py — SQLite connection helpers and query utilities.
"""

import sqlite3
import os
from pathlib import Path
from datetime import datetime, date
from typing import Optional, List, Dict, Any

DB_PATH = os.getenv("DATABASE_PATH", "./data/cfa_assistant.db")
SCHEMA_PATH = Path(__file__).parent / "schema.sql"


def is_postgres() -> bool:
    """Return True if a PostgreSQL/Supabase database URL is configured."""
    return bool(os.getenv("SUPABASE_DB_URL") or os.getenv("DATABASE_URL"))


class UnifiedCursor:
    """Wrapper around sqlite3.Cursor or psycopg2.cursor to unify placeholders and return dicts."""
    def __init__(self, cursor, is_pg: bool):
        self.cursor = cursor
        self.is_pg = is_pg

    def execute(self, query: str, params: tuple = ()) -> 'UnifiedCursor':
        if self.is_pg:
            # Convert SQLite '?' placeholders to PostgreSQL '%s'
            query_pg = query.replace('?', '%s')
            
            # Translate specific dialect queries
            if "INSERT OR IGNORE INTO curriculum_weights" in query_pg:
                query_pg = query_pg.replace(
                    "INSERT OR IGNORE INTO curriculum_weights (topic, weight) VALUES (%s, %s)",
                    "INSERT INTO curriculum_weights (topic, weight) VALUES (%s, %s) ON CONFLICT (topic) DO NOTHING"
                )
            elif "INSERT OR REPLACE INTO user_profiles" in query_pg:
                # user_profiles has user_id UNIQUE constraint
                query_pg = (
                    "INSERT INTO user_profiles (user_id, cfa_level, onboarding_done, phone) "
                    "VALUES (%s, %s, 0, %s) ON CONFLICT (user_id) DO UPDATE SET "
                    "cfa_level = EXCLUDED.cfa_level, phone = EXCLUDED.phone"
                )

            self.cursor.execute(query_pg, params)
        else:
            self.cursor.execute(query, params)
        return self

    def fetchone(self) -> Optional[Any]:
        row = self.cursor.fetchone()
        if row is None:
            return None
        if self.is_pg:
            return dict(row)
        return row

    def fetchall(self) -> List[Any]:
        rows = self.cursor.fetchall()
        if self.is_pg:
            return [dict(r) for r in rows]
        return rows

    @property
    def lastrowid(self) -> Optional[int]:
        if self.is_pg:
            return None
        return self.cursor.lastrowid


class UnifiedConnection:
    """Wrapper around sqlite3.Connection or psycopg2.connection."""
    def __init__(self, conn, is_pg: bool, pool=None):
        self.conn = conn
        self.is_pg = is_pg
        self.pool = pool
        self._active_cursor = None

    def execute(self, query: str, params: tuple = ()) -> UnifiedCursor:
        if self.is_pg:
            import psycopg2.extras
            if not self._active_cursor or self._active_cursor.closed:
                self._active_cursor = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            u_cur = UnifiedCursor(self._active_cursor, True)
            u_cur.execute(query, params)
            return u_cur
        else:
            cursor = self.conn.execute(query, params)
            return UnifiedCursor(cursor, False)

    def cursor(self):
        if self.is_pg:
            import psycopg2.extras
            return self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        else:
            return self.conn.cursor()

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        if self.is_pg:
            if self.pool:
                self.pool.putconn(self.conn)
            else:
                self.conn.close()
        else:
            self.conn.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            if exc_type is not None:
                self.conn.rollback()
            else:
                self.conn.commit()
        except Exception:
            pass
        if self.is_pg and self._active_cursor and not self._active_cursor.closed:
            self._active_cursor.close()
        self.close()


_local_pool = None

def get_postgres_pool(url: str):
    """Get or create a cached connection pool for PostgreSQL."""
    try:
        import streamlit as st
        if st.runtime.exists():
            @st.cache_resource
            def _get_cached_pool(connection_url: str):
                import psycopg2.pool
                return psycopg2.pool.ThreadedConnectionPool(1, 20, connection_url)
            return _get_cached_pool(url)
    except Exception:
        pass

    global _local_pool
    if _local_pool is None:
        import psycopg2.pool
        _local_pool = psycopg2.pool.ThreadedConnectionPool(1, 5, url)
    return _local_pool


def get_connection() -> UnifiedConnection:
    """Return a UnifiedConnection wrapper for either PostgreSQL or SQLite."""
    if is_postgres():
        try:
            import psycopg2
            import psycopg2.extras
            import psycopg2.pool
        except ImportError:
            raise ImportError(
                "PostgreSQL connection URL is configured, but 'psycopg2' is not installed. "
                "Please run: pip install psycopg2-binary"
            )
        url = os.getenv("SUPABASE_DB_URL") or os.getenv("DATABASE_URL")
        pool = get_postgres_pool(url)
        conn = pool.getconn()
        return UnifiedConnection(conn, True, pool)
    else:
        Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return UnifiedConnection(conn, False)



def init_db() -> None:
    """Create all tables and apply any pending migrations."""
    if is_postgres():
        schema_path = Path(__file__).parent / "schema_pg.sql"
        schema = schema_path.read_text()
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(schema)

        # Seed default weights in PostgreSQL
        with get_connection() as conn:
            row = conn.execute("SELECT COUNT(*) as total FROM curriculum_weights").fetchone()
            count = row["total"] if row else 0
            if count == 0:
                TOPIC_WEIGHTS = {
                    "Ethical and Professional Standards": 17.5,
                    "Quantitative Methods": 9.0,
                    "Economics": 9.0,
                    "Financial Statement Analysis": 12.5,
                    "Corporate Finance": 9.0,
                    "Equities": 11.5,
                    "Fixed Income": 11.5,
                    "Derivatives": 6.5,
                    "Alternative Investments": 6.5,
                    "Portfolio Construction": 6.5
                }
                for topic, weight in TOPIC_WEIGHTS.items():
                    conn.execute(
                        "INSERT INTO curriculum_weights (topic, weight) VALUES (?, ?) ON CONFLICT (topic) DO NOTHING",
                        (topic, weight)
                    )
        return

    # Local SQLite Initialization
    with get_connection() as conn:
        schema = SCHEMA_PATH.read_text()
        conn.conn.executescript(schema)

        # ── Column migrations ────────────────────────────────────────────────
        # Add source column to questions if missing (legacy migration)
        try:
            conn.execute("ALTER TABLE questions ADD COLUMN source TEXT DEFAULT 'ai'")
        except sqlite3.OperationalError:
            pass

        # Add creator column to scheduled_sessions if missing
        try:
            conn.execute("ALTER TABLE scheduled_sessions ADD COLUMN creator TEXT DEFAULT 'system'")
        except sqlite3.OperationalError:
            pass

        # Add completed_session_id column to scheduled_sessions if missing
        try:
            conn.execute("ALTER TABLE scheduled_sessions ADD COLUMN completed_session_id INTEGER REFERENCES study_sessions(id)")
        except sqlite3.OperationalError:
            pass

        # Add phone column to users if missing
        try:
            conn.execute("ALTER TABLE users ADD COLUMN phone TEXT")
        except sqlite3.OperationalError:
            pass

        # Remove cfa_level column from users if it exists (schema cleanup)
        try:
            conn.execute("ALTER TABLE users DROP COLUMN cfa_level")
        except sqlite3.OperationalError:
            pass

        # Seed curriculum weights if empty
        try:
            row = conn.execute("SELECT COUNT(*) as total FROM curriculum_weights").fetchone()
            count = row["total"] if row else 0
            if count == 0:
                TOPIC_WEIGHTS = {
                    "Ethical and Professional Standards": 17.5,
                    "Quantitative Methods": 9.0,
                    "Economics": 9.0,
                    "Financial Statement Analysis": 12.5,
                    "Corporate Finance": 9.0,
                    "Equities": 11.5,
                    "Fixed Income": 11.5,
                    "Derivatives": 6.5,
                    "Alternative Investments": 6.5,
                    "Portfolio Construction": 6.5
                }
                for topic, weight in TOPIC_WEIGHTS.items():
                    conn.execute("INSERT OR IGNORE INTO curriculum_weights (topic, weight) VALUES (?, ?)", (topic, weight))
        except Exception:
            pass

        # ── Topic name migrations (legacy → official 2027 names) ─────────────
        TOPIC_RENAMES = [
            ("Ethics & Professional Standards",  "Ethical and Professional Standards"),
            ("Ethics and Professional Standards", "Ethical and Professional Standards"),
            ("Equity Investments",               "Equities"),
            ("Corporate Issuers",                "Corporate Finance"),
            ("Portfolio Management",             "Portfolio Construction"),
        ]
        tables_with_topic = ["study_sessions", "questions", "topic_performance", "scheduled_sessions"]
        for old, new in TOPIC_RENAMES:
            for table in tables_with_topic:
                try:
                    conn.execute(f"UPDATE {table} SET topic=? WHERE topic=?", (new, old))
                except Exception:
                    pass


# ── User helpers ────────────────────────────────────────────────

def create_user(username: str, hashed_password: str, email: str = "", cfa_level: int = 1, phone: str = "") -> Optional[int]:
    try:
        with get_connection() as conn:
            if is_postgres():
                row = conn.execute(
                    "INSERT INTO users (username, password, email, phone) VALUES (?, ?, ?, ?) RETURNING id",
                    (username, hashed_password, email, phone),
                ).fetchone()
                user_id = row["id"] if row else None
            else:
                cur = conn.execute(
                    "INSERT INTO users (username, password, email, phone) VALUES (?, ?, ?, ?)",
                    (username, hashed_password, email, phone),
                )
                user_id = cur.lastrowid
            
            # Initialize profile with cfa_level in user_profiles
            conn.execute(
                "INSERT OR REPLACE INTO user_profiles (user_id, cfa_level, onboarding_done, phone) VALUES (?, ?, 0, ?)",
                (user_id, cfa_level, phone),
            )
            return user_id
    except Exception as e:
        if "IntegrityError" in type(e).__name__:
            return None
        raise e


def get_user_by_username(username: str) -> Optional[Dict]:
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        return dict(row) if row else None


def get_user_by_id(user_id: int) -> Optional[Dict]:
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        return dict(row) if row else None


# ── Study Session helpers ────────────────────────────────────────

def create_session(user_id: int, topic: str, session_type: str) -> int:
    with get_connection() as conn:
        if is_postgres():
            row = conn.execute(
                "INSERT INTO study_sessions (user_id, topic, session_type) VALUES (?, ?, ?) RETURNING id",
                (user_id, topic, session_type),
            ).fetchone()
            return row["id"] if row else None
        else:
            cur = conn.execute(
                "INSERT INTO study_sessions (user_id, topic, session_type) VALUES (?, ?, ?)",
                (user_id, topic, session_type),
            )
            return cur.lastrowid


def complete_session(session_id: int, score: float, total_q: int, correct_q: int, duration_mins: float) -> None:
    with get_connection() as conn:
        conn.execute(
            """UPDATE study_sessions SET score=?, total_q=?, correct_q=?, duration_mins=?, completed=1
               WHERE id=?""",
            (score, total_q, correct_q, duration_mins, session_id),
        )

        # Auto-link this completed session to a matching pending scheduled session
        try:
            row = conn.execute(
                "SELECT user_id, topic, session_type FROM study_sessions WHERE id=?", (session_id,)
            ).fetchone()
            if row:
                user_id = row["user_id"]
                topic = row["topic"]
                sess_type = row["session_type"]
                type_map = {
                    "practice": "Practice",
                    "mock": "Mock Exam",
                    "review": "Review",
                    "mixed": "Practice"
                }
                mapped_type = type_map.get(sess_type.lower(), "Practice")

                # Find oldest matching pending task
                scheduled = conn.execute(
                    """SELECT id FROM scheduled_sessions
                       WHERE user_id=? AND topic=? AND session_type=? AND status='pending'
                       ORDER BY scheduled_date ASC LIMIT 1""",
                    (user_id, topic, mapped_type)
                ).fetchone()
                if scheduled:
                    conn.execute(
                        "UPDATE scheduled_sessions SET status='done', completed_session_id=? WHERE id=?",
                        (session_id, scheduled["id"])
                    )
        except Exception:
            pass


def get_user_sessions(user_id: int, limit: int = 50) -> List[Dict]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM study_sessions WHERE user_id=? ORDER BY started_at DESC LIMIT ?",
            (user_id, limit),
        ).fetchall()
        return [dict(r) for r in rows]


# ── Question helpers ─────────────────────────────────────────────

def save_question(topic: str, subtopic: str, difficulty: str, question_text: str,
                  option_a: str, option_b: str, option_c: str,
                  correct_answer: str, explanation: str, source: str = 'ai') -> int:
    with get_connection() as conn:
        if is_postgres():
            row = conn.execute(
                """INSERT INTO questions
                   (topic, subtopic, difficulty, question_text, option_a, option_b, option_c, correct_answer, explanation, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?) RETURNING id""",
                (topic, subtopic, difficulty, question_text, option_a, option_b, option_c, correct_answer, explanation, source),
            ).fetchone()
            return row["id"] if row else None
        else:
            cur = conn.execute(
                """INSERT INTO questions
                   (topic, subtopic, difficulty, question_text, option_a, option_b, option_c, correct_answer, explanation, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (topic, subtopic, difficulty, question_text, option_a, option_b, option_c, correct_answer, explanation, source),
            )
            return cur.lastrowid


def get_question(question_id: int) -> Optional[Dict]:
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM questions WHERE id=?", (question_id,)).fetchone()
        return dict(row) if row else None


# ── Answer helpers ───────────────────────────────────────────────

def save_answer(user_id: int, question_id: int, session_id: int,
                selected: str, is_correct: bool, time_secs: float) -> None:
    with get_connection() as conn:
        conn.execute(
            """INSERT INTO user_answers
               (user_id, question_id, session_id, selected_answer, is_correct, time_taken_secs)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (user_id, question_id, session_id, selected, int(is_correct), time_secs),
        )


# ── Chat History helpers ─────────────────────────────────────────

def save_chat_message(user_id: int, role: str, content: str) -> None:
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO chat_history (user_id, role, content) VALUES (?, ?, ?)",
            (user_id, role, content),
        )


def get_chat_history(user_id: int, limit: int = 40) -> List[Dict]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT role, content FROM chat_history WHERE user_id=? ORDER BY timestamp DESC LIMIT ?",
            (user_id, limit),
        ).fetchall()
        return [dict(r) for r in reversed(rows)]


def clear_chat_history(user_id: int) -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM chat_history WHERE user_id=?", (user_id,))


def clear_all_user_data(user_id: int) -> None:
    """Delete ALL learning progress for a user (keeps account & profile but resets scores/history)."""
    with get_connection() as conn:
        conn.execute("DELETE FROM study_sessions       WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM user_answers         WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM chat_history         WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM scheduled_sessions   WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM topic_performance    WHERE user_id=?", (user_id,))


# ── Topic Performance helpers ─────────────────────────────────────

def upsert_topic_performance(user_id: int, topic: str, score: float) -> None:
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT avg_score, sessions_done FROM topic_performance WHERE user_id=? AND topic=?",
            (user_id, topic),
        ).fetchone()
        if existing:
            new_sessions = existing["sessions_done"] + 1
            new_avg = (existing["avg_score"] * existing["sessions_done"] + score) / new_sessions
            conn.execute(
                """UPDATE topic_performance SET avg_score=?, sessions_done=?, last_studied=CURRENT_TIMESTAMP
                   WHERE user_id=? AND topic=?""",
                (new_avg, new_sessions, user_id, topic),
            )
        else:
            conn.execute(
                """INSERT INTO topic_performance (user_id, topic, avg_score, sessions_done, last_studied)
                   VALUES (?, ?, ?, 1, CURRENT_TIMESTAMP)""",
                (user_id, topic, score),
            )


def get_topic_performance(user_id: int) -> List[Dict]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT topic, avg_score, sessions_done, last_studied FROM topic_performance WHERE user_id=?",
            (user_id,),
        ).fetchall()
        return [dict(r) for r in rows]


# ── Scheduled Sessions helpers ────────────────────────────────────

def save_scheduled_sessions(user_id: int, sessions: List[Dict], creator: str = "system") -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM scheduled_sessions WHERE user_id=? AND status='pending' AND creator=?", (user_id, creator))
        for s in sessions:
            conn.execute(
                """INSERT INTO scheduled_sessions (user_id, scheduled_date, topic, session_type, reason, priority, creator)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (user_id, s["date"], s["topic"], s["session_type"], s.get("reason", ""), s.get("priority", "medium"), creator),
            )


def get_upcoming_sessions(user_id: int, creator: Optional[str] = None) -> List[Dict]:
    today = date.today().isoformat()
    query = "SELECT * FROM scheduled_sessions WHERE user_id=? AND scheduled_date >= ? AND status='pending'"
    params = [user_id, today]
    if creator:
        query += " AND creator=?"
        params.append(creator)
    query += " ORDER BY scheduled_date ASC"
    
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


def mark_session_done(session_id: int, completed_session_id: Optional[int] = None) -> None:
    with get_connection() as conn:
        conn.execute(
            "UPDATE scheduled_sessions SET status='done', completed_session_id=? WHERE id=?",
            (completed_session_id, session_id),
        )


def add_manual_session(user_id: int, scheduled_date: str, topic: str, session_type: str, priority: str, reason: str) -> int:
    with get_connection() as conn:
        if is_postgres():
            row = conn.execute(
                """INSERT INTO scheduled_sessions (user_id, scheduled_date, topic, session_type, priority, reason, creator)
                   VALUES (?, ?, ?, ?, ?, ?, 'user') RETURNING id""",
                (user_id, scheduled_date, topic, session_type, priority, reason),
            ).fetchone()
            return row["id"] if row else None
        else:
            cur = conn.execute(
                """INSERT INTO scheduled_sessions (user_id, scheduled_date, topic, session_type, priority, reason, creator)
                   VALUES (?, ?, ?, ?, ?, ?, 'user')""",
                (user_id, scheduled_date, topic, session_type, priority, reason),
            )
            return cur.lastrowid


def delete_scheduled_session(session_id: int) -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM scheduled_sessions WHERE id=?", (session_id,))


# ── Question Bank helpers ────────────────────────────────────────

def get_bank_questions(topic: str, subtopics: Optional[List[str]] = None,
                       difficulty: Optional[str] = None, limit: int = 5) -> List[Dict[str, Any]]:
    """Retrieve questions from the imported question bank (source = 'bank')."""
    query = "SELECT * FROM questions WHERE source = 'bank' AND topic = ?"
    params = [topic]

    if subtopics:
        placeholders = ",".join("?" for _ in subtopics)
        query += f" AND subtopic IN ({placeholders})"
        params.extend(subtopics)

    if difficulty:
        query += " AND difficulty = ?"
        params.append(difficulty)

    query += " GROUP BY question_text ORDER BY RANDOM() LIMIT ?"
    params.append(limit)

    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


def delete_bank_questions(topic: Optional[str] = None) -> None:
    """Delete imported questions from the bank, optionally filtered by topic."""
    with get_connection() as conn:
        if topic:
            conn.execute("DELETE FROM questions WHERE source = 'bank' AND topic = ?", (topic,))
        else:
            conn.execute("DELETE FROM questions WHERE source = 'bank'")


def get_bank_stats() -> Dict[str, int]:
    """Get count of questions in bank grouped by topic."""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT topic, COUNT(*) as count FROM questions WHERE source = 'bank' GROUP BY topic"
        ).fetchall()
        return {r["topic"]: r["count"] for r in rows}


def seed_question_bank(xlsx_path: str) -> tuple[int, int]:
    """
    Read questions from the default Excel question bank and insert them into the DB.
    Deduplicates based on question text to avoid seeding duplicates.
    """
    import openpyxl
    from utils.cfa_topics import normalize_topic_name

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Question bank Excel file not found at {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Read header row (row 1) to map columns
    headers = [cell.value for cell in ws[1]]
    header_map = {}
    for idx, h in enumerate(headers):
        if h:
            header_map[h.strip().lower().replace(" ", "").replace("_", "")] = idx

    expected_headers = {
        "Topic": "topic",
        "Subtopic": "subtopic",
        "Difficulty": "difficulty",
        "Question": "question",
        "Option A": "option_a",
        "Option B": "option_b",
        "Option C": "option_c",
        "Correct Answer": "correct_answer",
        "Explanation": "explanation"
    }
    col_indices = {}
    for h_name, key in expected_headers.items():
        found = False
        for idx, h in enumerate(headers):
            if h and h.strip().lower() == h_name.lower():
                col_indices[key] = idx
                found = True
                break
        if not found:
            # Try fuzzy match
            clean_h_name = h_name.lower().replace(" ", "").replace("_", "")
            for idx, h in enumerate(headers):
                if h:
                    clean_h = str(h).strip().lower().replace(" ", "").replace("_", "")
                    if clean_h == clean_h_name:
                        col_indices[key] = idx
                        found = True
                        break
            if not found:
                raise ValueError(f"Missing column header in Excel: {h_name}")

    inserted = 0
    skipped = 0

    with get_connection() as conn:
        # Get existing questions in the bank to avoid duplicates
        existing_rows = conn.execute(
            "SELECT question_text FROM questions WHERE source = 'bank'"
        ).fetchall()
        existing_questions = {r["question_text"].strip().lower() for r in existing_rows}

        for row_idx in range(2, ws.max_row + 1):
            row_cells = [ws.cell(row_idx, col_idx + 1).value for col_idx in range(ws.max_column)]
            
            # Check if row is empty
            if not any(row_cells):
                continue
                
            def get_val(key):
                idx = col_indices.get(key)
                if idx is not None and idx < len(row_cells):
                    val = row_cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            topic = get_val("topic")
            subtopic = get_val("subtopic")
            difficulty = get_val("difficulty")
            question_text = get_val("question")
            option_a = get_val("option_a")
            option_b = get_val("option_b")
            option_c = get_val("option_c")
            correct_answer = get_val("correct_answer").upper()
            explanation = get_val("explanation")

            if not question_text:
                continue

            # Normalize topic name
            topic = normalize_topic_name(topic)
            
            # Normalize difficulty
            if difficulty.capitalize() in ["Easy", "Medium", "Hard"]:
                difficulty = difficulty.capitalize()
            else:
                difficulty = "Medium"

            if question_text.lower() in existing_questions:
                skipped += 1
                continue

            conn.execute(
                """INSERT INTO questions
                   (topic, subtopic, difficulty, question_text, option_a, option_b, option_c, correct_answer, explanation, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'bank')""",
                (topic, subtopic, difficulty, question_text, option_a, option_b, option_c, correct_answer, explanation),
            )
            existing_questions.add(question_text.lower())
            inserted += 1

    return inserted, skipped


# ── User Profile helpers ─────────────────────────────────────────────────────

def create_user_profile(
    user_id: int,
    full_name: str,
    age: int,
    gender: str,
    phone: str,
    city: str,
    cfa_level: int,
    exam_window: str,
    exam_year: int,
    exam_date: str,
) -> None:
    """Insert or replace a user profile row (upsert)."""
    with get_connection() as conn:
        conn.execute(
            """INSERT INTO user_profiles
               (user_id, full_name, age, gender, phone, city, cfa_level,
                exam_window, exam_year, exam_date, onboarding_done, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, CURRENT_TIMESTAMP)
               ON CONFLICT(user_id) DO UPDATE SET
                 full_name=excluded.full_name, age=excluded.age, gender=excluded.gender,
                 phone=excluded.phone, city=excluded.city, cfa_level=excluded.cfa_level,
                 exam_window=excluded.exam_window, exam_year=excluded.exam_year,
                 exam_date=excluded.exam_date, onboarding_done=1,
                 updated_at=CURRENT_TIMESTAMP""",
            (user_id, full_name, age, gender, phone, city, cfa_level,
             exam_window, exam_year, exam_date),
        )


def get_user_profile(user_id: int) -> Optional[Dict]:
    """Fetch the full profile for a user (returns None if not found)."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM user_profiles WHERE user_id=?", (user_id,)
        ).fetchone()
        return dict(row) if row else None


def update_user_profile(user_id: int, **kwargs) -> None:
    """Update arbitrary fields in the user profile."""
    if not kwargs:
        return
    fields = ", ".join(f"{k}=?" for k in kwargs)
    values = list(kwargs.values()) + [user_id]
    with get_connection() as conn:
        conn.execute(
            f"UPDATE user_profiles SET {fields}, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
            values,
        )


def is_onboarding_done(user_id: int) -> bool:
    """Return True if the user has completed onboarding."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT onboarding_done FROM user_profiles WHERE user_id=?", (user_id,)
        ).fetchone()
        return bool(row and row["onboarding_done"])


def get_subscription_status(user_id: int) -> Dict:
    """Return subscription info for the user (defaults to free/inactive)."""
    with get_connection() as conn:
        row = conn.execute(
            """SELECT subscription_plan, subscription_status, subscription_expires
               FROM user_profiles WHERE user_id=?""",
            (user_id,),
        ).fetchone()
        if row:
            return dict(row)
        return {
            "subscription_plan": "free",
            "subscription_status": "inactive",
            "subscription_expires": None,
        }


def get_subtopic_performance(user_id: int) -> List[Dict]:
    """Calculate subtopic performance dynamically from user_answers."""
    with get_connection() as conn:
        rows = conn.execute(
            """SELECT q.topic, q.subtopic, AVG(a.is_correct) * 100 as avg_score, COUNT(a.id) as total_answers
               FROM user_answers a
               JOIN questions q ON a.question_id = q.id
               WHERE a.user_id = ? AND q.subtopic IS NOT NULL AND q.subtopic != ''
               GROUP BY q.topic, q.subtopic
               ORDER BY avg_score ASC""",
            (user_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def get_curriculum_weights() -> Dict[str, float]:
    """Retrieve official CFA Level I weights from the database."""
    with get_connection() as conn:
        rows = conn.execute("SELECT topic, weight FROM curriculum_weights").fetchall()
        if not rows:
            from utils.cfa_topics import TOPIC_WEIGHTS
            return TOPIC_WEIGHTS
        return {r["topic"]: r["weight"] for r in rows}

